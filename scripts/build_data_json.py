import json
import math
import re
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
OUT_JSON = ROOT / "data.json"

CANDIDATE_XLSX = [
    ROOT / "CW _ CHOTO Weekly Expenditure Template 2.xlsx",
    ROOT / "CW _  CHOTO Weekly Expenditure Template 2.xlsx",
    ROOT / "source.xlsx",
]


def is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and math.isfinite(value)


def build_payload(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["CHOTO_WEEKLY EXPENDITURE"]

    week_cols = []
    weeks = []
    dates = []

    for c in range(1, ws.max_column + 1):
        w = ws.cell(3, c).value
        if isinstance(w, str) and re.fullmatch(r"Week\s+\d+", w.strip()):
            week_cols.append(c)
            weeks.append(w.strip())
            d = ws.cell(4, c).value
            dates.append(str(d).strip() if d is not None else "")

    rows = []
    for r in range(5, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if not isinstance(label, str):
            continue

        raw = label.rstrip()
        if not re.match(r"^\s*\d{4}", raw):
            continue

        vals = []
        numeric_count = 0
        for c in week_cols:
            v = ws.cell(r, c).value
            if is_number(v):
                vals.append(round(float(v), 2))
                numeric_count += 1
            else:
                vals.append(None)

        if numeric_count == 0:
            continue

        stripped = raw.strip()
        m = re.match(r"^(\d{4}(?:-\d{2,3})*)", stripped)
        code = m.group(1) if m else ""
        depth = len(code.split("-")) if code else 1
        name = re.sub(r"^\d{4}(?:-\d{2,3})*\s*[-–]?\s*", "", stripped).strip()

        rows.append(
            {
                "row": r,
                "label": stripped,
                "name": name or stripped,
                "code": code,
                "depth": depth,
                "values": vals,
            }
        )

    sales_target = []
    sales_actual = []
    for c in week_cols:
        t = ws.cell(1, c).value
        a = ws.cell(2, c).value
        sales_target.append(round(float(t), 2) if is_number(t) else 0.0)
        sales_actual.append(round(float(a), 2) if is_number(a) else 0.0)

    return {
        "generated_at": datetime.now(UTC).isoformat(),
        "data": {
            "weeks": weeks,
            "dates": dates,
            "rows": rows,
        },
        "sales_meta": {
            "target": sales_target,
            "actual": sales_actual,
        },
    }


def main():
    xlsx_path = next((p for p in CANDIDATE_XLSX if p.exists()), None)
    if not xlsx_path:
        checked = "\n".join(str(p) for p in CANDIDATE_XLSX)
        raise FileNotFoundError(f"Excel file not found. Checked:\n{checked}")

    print(f"Using Excel file: {xlsx_path}")
    payload = build_payload(xlsx_path)
    OUT_JSON.write_text(json.dumps(payload, separators=(",", ":")), encoding="utf-8")
    print(f"Wrote {OUT_JSON}")
    print(f"Rows: {len(payload['data']['rows'])}, Weeks: {len(payload['data']['weeks'])}")


if __name__ == "__main__":
    main()
