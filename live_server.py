import json
import math
import re
from http.server import HTTPServer, SimpleHTTPRequestHandler
from pathlib import Path
import openpyxl

BASE = Path(r"C:\Users\ukart\OneDrive - University of Tennessee\M\INtern\Part-time\choto-dashboard-vercel\choto-dashboard-vercel2")
XLSX = Path(r"C:\Users\ukart\OneDrive - University of Tennessee\M\INtern\Part-time\CW _  CHOTO Weekly Expenditure Template 2.xlsx")


def _num(v):
    return round(float(v), 2) if isinstance(v, (int, float)) and math.isfinite(v) else None


def _num0(v):
    return round(float(v), 2) if isinstance(v, (int, float)) and math.isfinite(v) else 0.0


def build_payload():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb['CHOTO_WEEKLY EXPENDITURE']

    week_cols, weeks, dates = [], [], []
    for c in range(1, ws.max_column + 1):
        w = ws.cell(3, c).value
        if isinstance(w, str) and re.fullmatch(r"Week\s+\d+", w.strip()):
            week_cols.append(c)
            weeks.append(w.strip())
            d = ws.cell(4, c).value
            dates.append(str(d).strip() if d is not None else "")

    rows = []
    for r in range(5, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if not isinstance(label, str):
            continue
        raw = label.rstrip()
        if not re.match(r"^\s*\d{4}", raw):
            continue

        vals = []
        numeric_count = 0
        for c in week_cols:
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)) and math.isfinite(v):
                vals.append(round(float(v), 2))
                numeric_count += 1
            else:
                vals.append(None)
        if numeric_count == 0:
            continue

        stripped = raw.strip()
        code_match = re.match(r"^(\d{4}(?:-\d{2,3})*)", stripped)
        code = code_match.group(1) if code_match else ""
        depth = len(code.split('-')) if code else 1
        name = re.sub(r"^\d{4}(?:-\d{2,3})*\s*[-?]?\s*", "", stripped).strip()

        rows.append({
            "row": r,
            "label": stripped,
            "name": name or stripped,
            "code": code,
            "depth": depth,
            "values": vals,
        })

    sales_meta = {
        "target": [_num0(ws.cell(1, c).value) for c in week_cols],
        "actual": [_num0(ws.cell(2, c).value) for c in week_cols],
    }

    return {
        "data": {"weeks": weeks, "dates": dates, "rows": rows},
        "sales_meta": sales_meta,
    }


class Handler(SimpleHTTPRequestHandler):
    def do_GET(self):
        if self.path.startswith('/data.json'):
            try:
                payload = build_payload()
                body = json.dumps(payload).encode('utf-8')
                self.send_response(200)
                self.send_header('Content-Type', 'application/json; charset=utf-8')
                self.send_header('Cache-Control', 'no-store, no-cache, must-revalidate, max-age=0')
                self.send_header('Content-Length', str(len(body)))
                self.end_headers()
                self.wfile.write(body)
            except Exception as e:
                msg = json.dumps({"error": str(e)}).encode('utf-8')
                self.send_response(500)
                self.send_header('Content-Type', 'application/json; charset=utf-8')
                self.send_header('Content-Length', str(len(msg)))
                self.end_headers()
                self.wfile.write(msg)
            return
        return super().do_GET()


if __name__ == '__main__':
    import os
    os.chdir(BASE)
    server = HTTPServer(('127.0.0.1', 8080), Handler)
    print('Serving live dashboard on http://127.0.0.1:8080')
    server.serve_forever()
